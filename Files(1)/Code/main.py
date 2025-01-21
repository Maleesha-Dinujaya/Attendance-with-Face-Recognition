import os
import pickle
import numpy as np
import cv2
import json
import face_recognition
import cvzone
from datetime import datetime
from openpyxl import load_workbook

# Paths
excel_file_path = "C:\\Users\\malee\\OneDrive\\Desktop\\Face Recognition Attendence System\\attendance.xlsx"
json_file_path = "students.json"
background_image_path = 'C:\\Users\\malee\\OneDrive\\Desktop\\Face Recognition Attendence System\\Files(1)\\Resources\\background.png'
mode_folder_path = 'C:\\Users\\malee\\OneDrive\\Desktop\\Face Recognition Attendence System\\Files(1)\\Resources\\Modes'
encoding_file_path = 'EncodeFile.p'

# Load student data from JSON file
try:
    with open(json_file_path, 'r') as f:
        data = json.load(f)
except FileNotFoundError:
    print("Error: JSON file not found.")
    data = {}

# Initialize camera
cap = cv2.VideoCapture(0)
cap.set(3, 640)
cap.set(4, 480)

# Load the background image
imgBackground = cv2.imread(background_image_path)
if imgBackground is None:
    print("Error: Could not load background image. Check the file path.")
    exit()

# Load mode images
mode_images = []
if os.path.exists(mode_folder_path):
    mode_files = os.listdir(mode_folder_path)
    mode_images = [cv2.imread(os.path.join(mode_folder_path, file)) for file in mode_files]
else:
    print("Error: Mode folder not found.")
    exit()

# Load encoding file
try:
    print("Loading Encode File...")
    with open(encoding_file_path, 'rb') as file:
        encode_list_with_ids = pickle.load(file)
    print("Encode File Loaded")
except FileNotFoundError:
    print("Error: Encoding file not found.")
    exit()

# Extract encodings and student IDs
encode_list_known = [entry["encoding"] for entry in encode_list_with_ids]
student_ids = [entry["id"] for entry in encode_list_with_ids]

mode_type = 0
counter = 0
current_id = -1
img_student = []

# Check if camera is accessible
if not cap.isOpened():
    print("Error: Could not access the camera.")
    exit()

while True:
    success, img = cap.read()
    if not success:
        print("Error: Failed to capture image.")
        break

    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    # Process face locations and encodings
    face_cur_frame = face_recognition.face_locations(imgS)
    encode_cur_frame = face_recognition.face_encodings(imgS, face_cur_frame)

    # Overlay captured image on background
    img_resized = cv2.resize(img, (640, 480))
    imgBackground[162:162 + 480, 55:55 + 640] = img_resized
    imgBackground[44:44 + 633, 808:808 + 414] = mode_images[0]

    for encode_face, face_loc in zip(encode_cur_frame, face_cur_frame):
        matches = face_recognition.compare_faces(encode_list_known, encode_face)
        face_distances = face_recognition.face_distance(encode_list_known, encode_face)

        if matches and np.any(matches):
            match_index = np.argmin(face_distances)
            matched_student_id = student_ids[match_index]

            # Retrieve student info from JSON
            student_info = data.get(str(matched_student_id))
            if student_info:
                print(f"Match Found: ID={matched_student_id}, Name={student_info['name']}")
            else:
                print(f"Error: No data found for ID {matched_student_id}.")
                continue

            # Draw bounding box
            y1, x2, y2, x1 = face_loc
            y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
            bbox = (55 + x1, 162 + y1, x2 - x1, y2 - y1)
            imgBackground = cvzone.cornerRect(imgBackground, bbox, rt=0)

            # Attendance update logic
            last_attendance_time = datetime.strptime(student_info['last_attendance_time'], "%Y-%m-%d %H:%M:%S")
            time_elapsed = (datetime.now() - last_attendance_time).total_seconds()

            if time_elapsed > 30:
                student_info['total_attendance'] += 1
                student_info['last_attendance_time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Save updated data to JSON
                with open(json_file_path, 'w') as f:
                    json.dump(data, f, indent=4)

                # Update Excel file
                try:
                    workbook = load_workbook(excel_file_path)
                    sheet = workbook.active
                    next_row = sheet.max_row + 1
                    sheet.cell(row=next_row, column=1, value=matched_student_id)
                    sheet.cell(row=next_row, column=2, value=student_info['name'])
                    sheet.cell(row=next_row, column=3, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    workbook.save(excel_file_path)
                    print(f"Attendance recorded for {student_info['name']} in Excel.")
                except FileNotFoundError:
                    print("Error: Excel file not found.")
            else:
                print(f"Attendance already recorded recently for {student_info['name']}.")
                mode_type = 3
                counter = 0
                imgBackground[44:44 + 633, 808:808 + 414] = mode_images[mode_type]

            # Display student details on the screen
            img_student = cv2.imread(student_info['image'])
            if img_student is not None:
                img_student = cv2.resize(img_student, (216, 216))
                imgBackground[175:175 + 216, 909:909 + 216] = img_student

            cv2.putText(imgBackground, str(student_info['name']), (900, 125), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 1)

    # Display the final output
    cv2.imshow("Face Attendance", imgBackground)
    key = cv2.waitKey(1)

    if key == 27:  # Exit on pressing the ESC key
        break

# Release camera and close windows
cap.release()
cv2.destroyAllWindows()